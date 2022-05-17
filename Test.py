from untitled import Tests,run
from openpyxl import Workbook,load_workbook

wb = Workbook()
wb.create_sheet("sheet_1",0)
wb.save('appText.xlsx')
wb.close()
wb = load_workbook('./appText.xlsx')
ws = wb.active
ws.cell(row=1, column=2, value='input')
ws.cell(row=1, column=3, value='output')
wb.save('appText.xlsx')
wb.close()
i =2
while True:
   print("open ?")
   a,b = run()
   ws.cell(row=i, column=2, value=a)
   ws.cell(row=i, column=3, value=b)
   print(b)
   i = i+1
   wb.save('appText.xlsx')
   wb.close()
x = 0
for x in range(100):
   openVideoApp()
   openMusicApp()

   x = x+1
print("finish text")